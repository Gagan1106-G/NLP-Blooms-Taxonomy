"""Export utilities for multi-type question sets.

Provides functions to export generated questions (MCQ, True/False, Short
Answer) to CSV, Excel, PDF, and pandas DataFrame formats.  All
file-producing functions return a ``BytesIO`` buffer so they integrate
directly with Streamlit's ``st.download_button``.

Supported question types
------------------------
* **MCQ** – question, options (A–D), answer label, answer, Bloom level,
  difficulty.
* **True/False** – question, boolean answer, optional explanation, Bloom
  level, difficulty.
* **Short Answer** – fill-in-the-blank question, answer text, keyword
  hints, Bloom level, difficulty.
"""

from __future__ import annotations

import csv
import io
import logging
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    HRFlowable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------
_OPTION_LABELS = ["A", "B", "C", "D"]


def _validate_mcqs(mcqs: list[dict[str, Any]], caller: str) -> None:
    """Raise ``ValueError`` if *mcqs* is empty or malformed."""
    if not mcqs:
        raise ValueError(f"{caller}: mcqs list must not be empty.")


def _option_columns(mcq: dict[str, Any]) -> dict[str, str]:
    """Return a mapping ``{Option A: text, Option B: text, …}`` for one MCQ."""
    options: list[str] = mcq.get("options", [])
    return {
        f"Option {label}": options[i] if i < len(options) else ""
        for i, label in enumerate(_OPTION_LABELS)
    }


def _answer_label(mcq: dict[str, Any]) -> str:
    """Return the answer label (A–D) from the MCQ dict."""
    label = mcq.get("answer_label", "")
    if label:
        return label
    # Fallback: derive label from answer text
    answer = mcq.get("answer", "")
    options = mcq.get("options", [])
    for i, opt in enumerate(options):
        if opt == answer and i < len(_OPTION_LABELS):
            return _OPTION_LABELS[i]
    return answer


# ---------------------------------------------------------------------------
# Type-detection helpers
# ---------------------------------------------------------------------------
def _qtype(q: dict[str, Any]) -> str:
    """Return normalised question type: 'MCQ', 'True/False', or 'Short Answer'."""
    return q.get("type", "MCQ")


def _is_mcq(q: dict[str, Any]) -> bool:
    return _qtype(q) == "MCQ"


def _is_tf(q: dict[str, Any]) -> bool:
    return _qtype(q) == "True/False"


def _is_sa(q: dict[str, Any]) -> bool:
    return _qtype(q) == "Short Answer"


# ---------------------------------------------------------------------------
# 1. DataFrame conversion
# ---------------------------------------------------------------------------
def mcqs_to_dataframe(questions: list[dict[str, Any]]) -> pd.DataFrame:
    """Convert a list of question dictionaries to a :class:`pandas.DataFrame`.

    Handles all three question types (MCQ, True/False, Short Answer) in a
    unified schema.  Columns that are not applicable to a given type are
    filled with an empty string.

    Columns
    -------
    ``No``, ``Type``, ``Question``,
    ``Option A``, ``Option B``, ``Option C``, ``Option D``,
    ``Answer Label``, ``Answer``, ``Keywords``, ``Explanation``,
    ``Bloom Level``, ``Difficulty``.

    Args:
        questions: List of question dicts (any mix of types).

    Returns:
        A DataFrame with the columns described above.

    Raises:
        ValueError: If *questions* is empty.
    """
    _validate_mcqs(questions, "mcqs_to_dataframe")

    rows: list[dict[str, Any]] = []
    for i, q in enumerate(questions, start=1):
        qt = _qtype(q)
        row: dict[str, Any] = {
            "No":           i,
            "Type":         qt,
            "Question":     q.get("question", ""),
            "Option A":     "",
            "Option B":     "",
            "Option C":     "",
            "Option D":     "",
            "Answer Label": "",
            "Answer":       "",
            "Keywords":     "",
            "Explanation":  "",
            "Bloom Level":  q.get("level", ""),
            "Difficulty":   q.get("difficulty", ""),
        }

        if _is_mcq(q):
            opts = q.get("options", [])
            for j, label in enumerate(_OPTION_LABELS):
                row[f"Option {label}"] = opts[j] if j < len(opts) else ""
            row["Answer Label"] = _answer_label(q)
            row["Answer"]       = q.get("answer", "")

        elif _is_tf(q):
            raw_ans = q.get("answer", True)
            row["Answer"]      = "True" if raw_ans else "False"
            row["Explanation"] = q.get("explanation", "")

        elif _is_sa(q):
            row["Answer"]   = q.get("answer", "")
            kws = q.get("keywords", [])
            row["Keywords"] = ", ".join(kws) if isinstance(kws, list) else str(kws)

        rows.append(row)

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# 2. CSV export
# ---------------------------------------------------------------------------
def export_to_csv(
    mcqs: list[dict[str, Any]],
    filename: str = "questions.csv",  # noqa: ARG001 – kept for API consistency
) -> io.BytesIO:
    """Export questions to a CSV file returned as a ``BytesIO`` buffer.

    Handles MCQ, True/False, and Short Answer types in a unified schema.

    Columns: ``No``, ``Type``, ``Question``, ``Option A–D``,
    ``Answer Label``, ``Answer``, ``Keywords``, ``Explanation``,
    ``Bloom Level``, ``Difficulty``.

    Args:
        mcqs: List of question dictionaries (any mix of types).
        filename: Logical filename (unused internally; preserved for API
            consistency with other export functions).

    Returns:
        A ``BytesIO`` buffer containing UTF-8 encoded CSV data.

    Raises:
        ValueError: If *mcqs* is empty.
        RuntimeError: On unexpected serialisation errors.

    Example:
        >>> buf = export_to_csv(questions, "quiz.csv")
        >>> st.download_button("Download CSV", buf, "quiz.csv", "text/csv")
    """
    _validate_mcqs(mcqs, "export_to_csv")

    try:
        df = mcqs_to_dataframe(mcqs)
        buffer = io.BytesIO()
        df.to_csv(buffer, index=False, encoding="utf-8-sig")  # BOM for Excel compat
        buffer.seek(0)
        logger.info("CSV export: %d question(s) written.", len(mcqs))
        return buffer
    except Exception as exc:
        logger.error("export_to_csv failed: %s", exc)
        raise RuntimeError(f"CSV export failed: {exc}") from exc


# ---------------------------------------------------------------------------
# 3. Excel export
# ---------------------------------------------------------------------------
_HEADER_FILL   = PatternFill("solid", fgColor="2E4057")
_SUBHEAD_FILL  = PatternFill("solid", fgColor="048A81")
_ALT_ROW_FILL  = PatternFill("solid", fgColor="F0F4F8")

_THIN_BORDER   = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_ANSWER_FILL   = PatternFill("solid", fgColor="D4EDDA")

# Per-type accent fills (left-border colour is applied via cell fill on col 2)
_TYPE_ROW_FILLS: dict[str, PatternFill] = {
    "MCQ":          PatternFill("solid", fgColor="E3F2FD"),   # light blue
    "True/False":   PatternFill("solid", fgColor="E8F5E9"),   # light green
    "Short Answer": PatternFill("solid", fgColor="FCE4EC"),   # light pink
}
_TYPE_BADGE_FILLS: dict[str, PatternFill] = {
    "MCQ":          PatternFill("solid", fgColor="1565C0"),
    "True/False":   PatternFill("solid", fgColor="2E7D32"),
    "Short Answer": PatternFill("solid", fgColor="AD1457"),
}


def export_to_excel(
    mcqs: list[dict[str, Any]],
    filename: str = "questions.xlsx",  # noqa: ARG001
    sheet_title: str = "Question Bank",
) -> io.BytesIO:
    """Export questions to a formatted Excel workbook returned as ``BytesIO``.

    Handles MCQ, True/False, and Short Answer types.  Each type is
    colour-coded in the **Type** column.  Produces two sheets:

    * **Question Bank** – main data table with type-coloured rows.
    * **Summary** – per-type, per-Bloom-level, and per-difficulty counts.

    Args:
        mcqs: List of question dictionaries (any mix of types).
        filename: Logical filename (unused internally).
        sheet_title: Title for the main worksheet.

    Returns:
        A ``BytesIO`` buffer containing the ``.xlsx`` file.

    Raises:
        ValueError: If *mcqs* is empty.
        RuntimeError: On unexpected serialisation errors.

    Example:
        >>> buf = export_to_excel(questions, "quiz.xlsx")
        >>> st.download_button("Download Excel", buf, "quiz.xlsx",
        ...     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    """
    _validate_mcqs(mcqs, "export_to_excel")

    try:
        wb = Workbook()

        # ── Main sheet ───────────────────────────────────────────────────────
        ws = wb.active
        ws.title = sheet_title

        columns = [
            ("No",           5),
            ("Type",         14),
            ("Question",     50),
            ("Option A",     24),
            ("Option B",     24),
            ("Option C",     24),
            ("Option D",     24),
            ("Answer Label", 13),
            ("Answer",       26),
            ("Keywords",     22),
            ("Explanation",  28),
            ("Bloom Level",  16),
            ("Difficulty",   12),
        ]

        # Title row
        ws.merge_cells(f"A1:{get_column_letter(len(columns))}1")
        title_cell = ws["A1"]
        title_cell.value = f"Question Bank  –  Generated {datetime.now():%d %b %Y}"
        title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        title_cell.fill = _HEADER_FILL
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Header row
        header_row = 2
        col_index: dict[str, int] = {}
        for col_idx, (col_name, col_width) in enumerate(columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            cell.fill = _SUBHEAD_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width
            col_index[col_name] = col_idx
        ws.row_dimensions[header_row].height = 22

        # Data rows
        df = mcqs_to_dataframe(mcqs)
        type_col_idx   = col_index["Type"]
        ans_lbl_col    = col_index["Answer Label"]

        for row_offset, (_, row) in enumerate(df.iterrows()):
            excel_row = header_row + 1 + row_offset
            qt = str(row.get("Type", "MCQ"))

            # Base fill: alternating white / light-grey; override with type shade
            base_fill = _ALT_ROW_FILL if row_offset % 2 == 1 else None

            for col_idx, (col_name, _) in enumerate(columns, start=1):
                value = row.get(col_name, "")
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )
                cell.border = _THIN_BORDER
                if base_fill:
                    cell.fill = base_fill

            # Type badge cell – distinct background + white bold text
            type_cell = ws.cell(row=excel_row, column=type_col_idx)
            type_cell.fill = _TYPE_BADGE_FILLS.get(qt, PatternFill("solid", fgColor="607D8B"))
            type_cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            type_cell.alignment = Alignment(horizontal="center", vertical="top")

            # Answer label cell – green highlight for MCQ only
            if qt == "MCQ":
                ans_cell = ws.cell(row=excel_row, column=ans_lbl_col)
                ans_cell.fill = _ANSWER_FILL
                ans_cell.font = Font(name="Calibri", bold=True, size=10, color="155724")
                ans_cell.alignment = Alignment(horizontal="center", vertical="top")

            ws.row_dimensions[excel_row].height = 40

        ws.freeze_panes = f"A{header_row + 1}"

        # ── Summary sheet ────────────────────────────────────────────────────
        ws2 = wb.create_sheet("Summary")
        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 12

        def _write_section_header(row: int, heading: str) -> None:
            ws2.cell(row=row, column=1, value=heading).font = Font(bold=True, size=11, color="FFFFFF")
            ws2.cell(row=row, column=1).fill = _SUBHEAD_FILL
            ws2.cell(row=row, column=2, value="Count").font = Font(bold=True, color="FFFFFF")
            ws2.cell(row=row, column=2).fill = _SUBHEAD_FILL

        # Title
        ws2["A1"] = "Export Summary"
        ws2["A1"].font = Font(bold=True, size=13, color="FFFFFF")
        ws2["A1"].fill = _HEADER_FILL

        cur_row = 3
        # Question type counts
        _write_section_header(cur_row, "Question Type")
        cur_row += 1
        type_counts = df["Type"].value_counts()
        for t, cnt in type_counts.items():
            ws2.cell(row=cur_row, column=1, value=t).border  = _THIN_BORDER
            ws2.cell(row=cur_row, column=2, value=int(cnt)).border = _THIN_BORDER
            ws2.cell(row=cur_row, column=1).fill = _TYPE_BADGE_FILLS.get(
                str(t), PatternFill("solid", fgColor="607D8B")
            )
            ws2.cell(row=cur_row, column=1).font = Font(color="FFFFFF", bold=True)
            cur_row += 1

        cur_row += 1
        # Bloom level counts
        _write_section_header(cur_row, "Bloom Level")
        cur_row += 1
        bloom_counts = df["Bloom Level"].replace("", pd.NA).dropna().value_counts()
        for level, cnt in bloom_counts.items():
            ws2.cell(row=cur_row, column=1, value=level).border = _THIN_BORDER
            ws2.cell(row=cur_row, column=2, value=int(cnt)).border = _THIN_BORDER
            cur_row += 1

        cur_row += 1
        # Difficulty counts
        _write_section_header(cur_row, "Difficulty")
        cur_row += 1
        diff_counts = df["Difficulty"].replace("", pd.NA).dropna().value_counts()
        for diff, cnt in diff_counts.items():
            ws2.cell(row=cur_row, column=1, value=diff).border = _THIN_BORDER
            ws2.cell(row=cur_row, column=2, value=int(cnt)).border = _THIN_BORDER
            cur_row += 1

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        logger.info("Excel export: %d question(s) written.", len(mcqs))
        return buffer

    except Exception as exc:
        logger.error("export_to_excel failed: %s", exc)
        raise RuntimeError(f"Excel export failed: {exc}") from exc


# ---------------------------------------------------------------------------
# 4. PDF export
# ---------------------------------------------------------------------------
def export_to_pdf(
    mcqs: list[dict[str, Any]],
    filename: str = "questions.pdf",  # noqa: ARG001
    title: str = "Question Paper",
    institution: str = "",
    show_answers: bool = False,
) -> io.BytesIO:
    """Export questions to a formatted PDF test paper returned as ``BytesIO``.

    Handles MCQ, True/False, and Short Answer types with distinct visual
    formatting per type:

    * **MCQ** – numbered question + four lettered options (A–D).
    * **True/False** – question + "☐ True  ☐ False" tick boxes.
    * **Short Answer** – question stem (with blank) + keyword hints if present.

    When *show_answers* is ``True`` an answer key section is appended,
    formatted appropriately for each question type.

    Args:
        mcqs: List of question dictionaries (any mix of types).
        filename: Logical filename (unused internally).
        title: Heading printed at the top of the document.
        institution: Optional institution/course name shown beneath the title.
        show_answers: When ``True``, append an answer key section.

    Returns:
        A ``BytesIO`` buffer containing the PDF.

    Raises:
        ValueError: If *mcqs* is empty.
        RuntimeError: On unexpected serialisation errors.

    Example:
        >>> buf = export_to_pdf(questions, title="Biology Quiz", show_answers=True)
        >>> st.download_button("Download PDF", buf, "quiz.pdf", "application/pdf")
    """
    _validate_mcqs(mcqs, "export_to_pdf")

    # Accent colours per question type
    _TYPE_ACCENT: dict[str, colors.Color] = {
        "MCQ":          colors.HexColor("#1565C0"),
        "True/False":   colors.HexColor("#2E7D32"),
        "Short Answer": colors.HexColor("#AD1457"),
    }

    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=2 * cm,
            rightMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )

        # ── Styles ───────────────────────────────────────────────────────────
        base_styles = getSampleStyleSheet()

        style_title = ParagraphStyle(
            "QuizTitle",
            parent=base_styles["Title"],
            fontSize=18,
            leading=22,
            textColor=colors.HexColor("#2E4057"),
            alignment=TA_CENTER,
            spaceAfter=4,
        )
        style_subtitle = ParagraphStyle(
            "QuizSubtitle",
            fontSize=11,
            leading=14,
            textColor=colors.HexColor("#555555"),
            alignment=TA_CENTER,
            spaceAfter=2,
        )
        style_meta = ParagraphStyle(
            "QuizMeta",
            fontSize=10,
            leading=13,
            textColor=colors.HexColor("#777777"),
            alignment=TA_CENTER,
            spaceAfter=12,
        )
        style_instructions = ParagraphStyle(
            "Instructions",
            fontSize=9,
            leading=13,
            textColor=colors.HexColor("#333333"),
            spaceAfter=8,
        )
        style_type_label = ParagraphStyle(
            "TypeLabel",
            fontSize=8,
            leading=10,
            fontName="Helvetica-Bold",
            textColor=colors.white,
            spaceAfter=3,
        )
        style_q = ParagraphStyle(
            "Question",
            fontSize=11,
            leading=15,
            textColor=colors.black,
            leftIndent=0,
            spaceAfter=4,
            fontName="Helvetica-Bold",
        )
        style_opt = ParagraphStyle(
            "Option",
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#222222"),
            leftIndent=18,
            spaceAfter=2,
        )
        style_tf_opt = ParagraphStyle(
            "TFOption",
            fontSize=11,
            leading=14,
            textColor=colors.HexColor("#222222"),
            leftIndent=18,
            spaceAfter=2,
        )
        style_sa_blank = ParagraphStyle(
            "SABlank",
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#555555"),
            leftIndent=18,
            spaceAfter=2,
            fontName="Helvetica-Oblique",
        )
        style_bloom = ParagraphStyle(
            "BloomTag",
            fontSize=8,
            leading=11,
            textColor=colors.HexColor("#048A81"),
            leftIndent=0,
            spaceAfter=10,
        )
        style_ans_head = ParagraphStyle(
            "AnswerHead",
            fontSize=13,
            leading=16,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#2E4057"),
            alignment=TA_LEFT,
            spaceBefore=8,
            spaceAfter=6,
        )
        style_key_note = ParagraphStyle(
            "KeyNote",
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#555555"),
            leftIndent=8,
            spaceAfter=4,
        )

        # ── Instruction text per type ─────────────────────────────────────────
        type_instruction_map = {
            "MCQ":          "circle the best answer (A–D)",
            "True/False":   "tick ☑ True or ☑ False",
            "Short Answer": "write your answer in the blank provided",
        }

        # ── Build content ────────────────────────────────────────────────────
        story: list[Any] = []

        # Header
        story.append(Paragraph(title, style_title))
        if institution:
            story.append(Paragraph(institution, style_subtitle))

        type_counts_summary = {}
        for q in mcqs:
            type_counts_summary[_qtype(q)] = type_counts_summary.get(_qtype(q), 0) + 1
        type_summary_str = "  |  ".join(
            f"{cnt} {t}" for t, cnt in type_counts_summary.items()
        )
        story.append(
            Paragraph(
                f"Date: {datetime.now():%d %B %Y} &nbsp;&nbsp;|&nbsp;&nbsp; "
                f"Total Questions: {len(mcqs)} ({type_summary_str}) "
                f"&nbsp;&nbsp;|&nbsp;&nbsp; Total Marks: {len(mcqs)}",
                style_meta,
            )
        )
        story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#2E4057")))
        story.append(Spacer(1, 0.3 * cm))

        # Per-type instructions
        active_types = list(type_counts_summary.keys())
        instructions_text = "  <b>·</b>  ".join(
            f"<b>{t}:</b> {type_instruction_map.get(t, 'answer each question')}"
            for t in active_types
        )
        story.append(
            Paragraph(
                f"<b>Instructions:</b> &nbsp;{instructions_text}. "
                "Each question carries <b>1 mark</b>.",
                style_instructions,
            )
        )
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#CCCCCC")))
        story.append(Spacer(1, 0.4 * cm))

        # ── Questions ────────────────────────────────────────────────────────
        for i, q in enumerate(mcqs, start=1):
            qt          = _qtype(q)
            accent      = _TYPE_ACCENT.get(qt, colors.HexColor("#607D8B"))
            q_text      = q.get("question", f"Question {i}")
            bloom_level = q.get("level", "")
            difficulty  = q.get("difficulty", "")

            # Type label badge (small coloured tag)
            badge_data = [[Paragraph(f" {qt} ", style_type_label)]]
            badge_table = Table(badge_data, colWidths=[2.4 * cm])
            badge_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), accent),
                ("TOPPADDING",    (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("LEFTPADDING",   (0, 0), (-1, -1), 4),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
                ("ROUNDEDCORNERS", [3]),
            ]))
            story.append(badge_table)
            story.append(Spacer(1, 0.1 * cm))

            # Question text
            story.append(Paragraph(f"<b>Q{i}.</b>  {q_text}", style_q))

            if _is_mcq(q):
                options = q.get("options", [])
                for j, label in enumerate(_OPTION_LABELS):
                    opt_text = options[j] if j < len(options) else "—"
                    story.append(Paragraph(f"({label})  {opt_text}", style_opt))

            elif _is_tf(q):
                story.append(Paragraph("☐  <b>True</b>", style_tf_opt))
                story.append(Paragraph("☐  <b>False</b>", style_tf_opt))

            elif _is_sa(q):
                story.append(
                    Paragraph(
                        "Answer: _________________________________________________",
                        style_sa_blank,
                    )
                )
                kws = q.get("keywords", [])
                if kws and isinstance(kws, list) and len(kws) > 0:
                    hints = ", ".join(kws[:4])
                    story.append(
                        Paragraph(f"<i>Hint keywords: {hints}</i>", style_sa_blank)
                    )

            # Bloom / difficulty meta tag
            meta_parts: list[str] = []
            if bloom_level:
                meta_parts.append(f"Bloom: {bloom_level}")
            if difficulty:
                meta_parts.append(f"Difficulty: {difficulty}")
            if meta_parts:
                story.append(Paragraph("  |  ".join(meta_parts), style_bloom))
            else:
                story.append(Spacer(1, 0.35 * cm))

        # ── Answer key ────────────────────────────────────────────────────────
        if show_answers:
            story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#2E4057")))
            story.append(Spacer(1, 0.3 * cm))
            story.append(Paragraph("Answer Key", style_ans_head))

            key_data: list[list[Any]] = []
            key_header = [
                Paragraph("<b>No.</b>",         base_styles["Normal"]),
                Paragraph("<b>Type</b>",         base_styles["Normal"]),
                Paragraph("<b>Answer</b>",        base_styles["Normal"]),
                Paragraph("<b>Bloom Level</b>",   base_styles["Normal"]),
                Paragraph("<b>Difficulty</b>",    base_styles["Normal"]),
                Paragraph("<b>Notes</b>",          base_styles["Normal"]),
            ]
            key_data.append(key_header)

            for i, q in enumerate(mcqs, start=1):
                qt    = _qtype(q)
                notes = ""

                if _is_mcq(q):
                    answer_text = f"({_answer_label(q)}) {q.get('answer', '')}"
                elif _is_tf(q):
                    raw = q.get("answer", True)
                    answer_text = "True" if raw else "False"
                    notes = q.get("explanation", "")[:80]   # truncate for table
                else:  # Short Answer
                    answer_text = q.get("answer", "")
                    kws = q.get("keywords", [])
                    if isinstance(kws, list) and kws:
                        notes = "KW: " + ", ".join(kws[:3])

                key_data.append([
                    str(i),
                    qt,
                    Paragraph(answer_text, base_styles["Normal"]),
                    q.get("level", ""),
                    q.get("difficulty", ""),
                    Paragraph(notes, base_styles["Normal"]),
                ])

            col_widths = [1.2 * cm, 2.6 * cm, None, 3.2 * cm, 2.4 * cm, 4.0 * cm]
            key_table = Table(key_data, colWidths=col_widths, repeatRows=1)

            # Base style
            ts = [
                ("BACKGROUND",     (0, 0), (-1, 0), colors.HexColor("#2E4057")),
                ("TEXTCOLOR",      (0, 0), (-1, 0), colors.white),
                ("FONTNAME",       (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",       (0, 0), (-1, -1), 9),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.white, colors.HexColor("#F0F4F8")]),
                ("GRID",           (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
                ("ALIGN",          (0, 0), (1, -1), "CENTER"),
                ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING",     (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING",  (0, 0), (-1, -1), 4),
            ]

            # Per-type colour on the Type column
            for row_idx, q in enumerate(mcqs, start=1):
                qt   = _qtype(q)
                fill = _TYPE_ACCENT.get(qt, colors.HexColor("#607D8B"))
                ts.append(("BACKGROUND", (1, row_idx), (1, row_idx), fill))
                ts.append(("TEXTCOLOR",  (1, row_idx), (1, row_idx), colors.white))
                ts.append(("FONTNAME",   (1, row_idx), (1, row_idx), "Helvetica-Bold"))

            key_table.setStyle(TableStyle(ts))
            story.append(key_table)

        doc.build(story)
        buffer.seek(0)
        logger.info(
            "PDF export: %d question(s) written (show_answers=%s).", len(mcqs), show_answers
        )
        return buffer

    except Exception as exc:
        logger.error("export_to_pdf failed: %s", exc)
        raise RuntimeError(f"PDF export failed: {exc}") from exc


# ---------------------------------------------------------------------------
# Example usage
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import pathlib

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s – %(message)s",
    )

    SAMPLE_QUESTIONS: list[dict[str, Any]] = [
        # MCQ
        {
            "type":         "MCQ",
            "question":     "What is the primary function of mitochondria?",
            "options":      ["Energy production", "Protein synthesis", "Cell division", "DNA storage"],
            "answer":       "Energy production",
            "answer_label": "A",
            "level":        "Remember",
            "difficulty":   "Easy",
        },
        {
            "type":         "MCQ",
            "question":     "Which process converts light energy into chemical energy?",
            "options":      ["Respiration", "Photosynthesis", "Fermentation", "Digestion"],
            "answer":       "Photosynthesis",
            "answer_label": "B",
            "level":        "Understand",
            "difficulty":   "Medium",
        },
        # True/False
        {
            "type":        "True/False",
            "question":    "Mitochondria are found in all eukaryotic cells.",
            "answer":      True,
            "explanation": "Eukaryotic cells universally contain mitochondria for energy production.",
            "level":       "Remember",
            "difficulty":  "Easy",
        },
        {
            "type":        "True/False",
            "question":    "Photosynthesis produces carbon dioxide as a by-product.",
            "answer":      False,
            "explanation": "Photosynthesis consumes CO₂ and produces oxygen.",
            "level":       "Understand",
            "difficulty":  "Easy",
        },
        # Short Answer
        {
            "type":       "Short Answer",
            "question":   "The powerhouse of the cell is the _____.",
            "answer":     "mitochondria",
            "keywords":   ["mitochondria", "organelle", "ATP"],
            "level":      "Remember",
            "difficulty": "Easy",
        },
        {
            "type":       "Short Answer",
            "question":   "The process by which plants make food using sunlight is called _____.",
            "answer":     "photosynthesis",
            "keywords":   ["photosynthesis", "chlorophyll", "light"],
            "level":      "Understand",
            "difficulty": "Easy",
        },
    ]

    out_dir = pathlib.Path("data/output")
    out_dir.mkdir(parents=True, exist_ok=True)

    # DataFrame
    df = mcqs_to_dataframe(SAMPLE_QUESTIONS)
    print("\n── DataFrame preview ──────────────────────────────────────")
    print(df.to_string(index=False))

    # CSV
    csv_buf = export_to_csv(SAMPLE_QUESTIONS)
    csv_path = out_dir / "sample_questions.csv"
    csv_path.write_bytes(csv_buf.read())
    print(f"\n✓ CSV  saved → {csv_path}")

    # Excel
    xl_buf  = export_to_excel(SAMPLE_QUESTIONS)
    xl_path = out_dir / "sample_questions.xlsx"
    xl_path.write_bytes(xl_buf.read())
    print(f"✓ Excel saved → {xl_path}")

    # PDF (without answers)
    pdf_buf  = export_to_pdf(SAMPLE_QUESTIONS, title="Mixed Question Paper",
                              institution="Demo School")
    pdf_path = out_dir / "sample_questions_paper.pdf"
    pdf_path.write_bytes(pdf_buf.read())
    print(f"✓ PDF  saved → {pdf_path}")

    # PDF with answer key
    pdf_ans_buf  = export_to_pdf(SAMPLE_QUESTIONS, title="Mixed Question Paper – Answer Key",
                                  institution="Demo School", show_answers=True)
    pdf_ans_path = out_dir / "sample_questions_answers.pdf"
    pdf_ans_path.write_bytes(pdf_ans_buf.read())
    print(f"✓ PDF (with answers) saved → {pdf_ans_path}")
